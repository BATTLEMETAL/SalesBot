import pandas as pd
from openpyxl import load_workbook

# NOTE: report_generator and chart_creator are imported only inside __main__
# to avoid circular dependencies (this module is imported by both of them).



def read_excel_file(file_path):
    """
    Reads an Excel file and returns a DataFrame containing sales data.
    
    Args:
        file_path (str): Path to the Excel file.
        
    Returns:
        pd.DataFrame: A DataFrame containing the sales data.
    """
    # Load the workbook
    wb = load_workbook(filename=file_path)
    
    # Select the first sheet
    sheet_name = wb.sheetnames[0]
    
    # Read the data into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    return df


def calculate_totals_by_region(df):
    """
    Calculates total sales per region.
    
    Args:
        df (pd.DataFrame): The input DataFrame containing sales data.
        
    Returns:
        pd.Series: A Series with regions as index and their corresponding total sales.
    """
    # Group by 'Region' and sum up the 'Sales'
    totals = df.groupby('Region')['Sales'].sum()
    
    return totals


def save_data_to_csv(df, output_path):
    """
    Saves the provided DataFrame to a CSV file.
    
    Args:
        df (pd.DataFrame): The DataFrame to be saved.
        output_path (str): The path where the CSV file will be saved.
    """
    # Save the DataFrame to a CSV file
    df.to_csv(output_path, index=False)


if __name__ == "__main__":
    from report_generator import generate_report
    from chart_creator import create_bar_chart

    # File paths
    input_file = "sales_data.xlsx"
    output_file = "processed_sales.csv"
    report_file = "sales_report.pdf"

    # Step 1: Read the Excel file
    print("Reading the Excel file...")
    data_frame = read_excel_file(input_file)

    # Step 2: Calculate totals by region
    print("Calculating totals by region...")
    totals = calculate_totals_by_region(data_frame)

    # Step 3: Save the processed data to a CSV file
    print(f"Saving processed data to {output_file}...")
    save_data_to_csv(totals, output_file)

    # Step 4: Generate the PDF report
    print(f"Generating the report at {report_file}...")
    totals_df = totals.reset_index()
    totals_df.columns = ["Region", "Total Sales"]
    generate_report(totals_df)

    # Step 5: Create a bar chart
    print("Creating a bar chart...")
    create_bar_chart(totals)

    print("Process completed successfully.")